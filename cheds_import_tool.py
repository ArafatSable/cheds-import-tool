#!/usr/bin/env python3
"""
cheds_import_tool.py
Reusable tool for turning raw Zoho-Creator import sheets into CHEDS/MOHESR
"import-ready" sheets, for ANY HEDB form (Employee-Workload, Institute-*,
Students-*, etc.) -- not just the one this was built for.

WHY THIS EXISTS
----------------
Each HEDB form's Zoho workflow script builds an API payload from `input.X`
fields. Column headers in a raw export rarely match Zoho's real field
"link names", and Lookup-type fields must be imported using the linked
report's descriptive Name text, not the numeric/short HEDB code (Zoho
matches lookups on Name, then the field's `.Code` sub-property is what
the workflow script sends to the API). This script automates that
translation using HEDB_.xlsx as the single source of truth for all
List-of-Values lookup tables.

THE FAST PATH FOR A NEW FORM (if you have the whole-app .ds export)
--------------------------------------------------------------------
0a) list-forms       -- list every form name inside a whole-application .ds
                         export (Zoho Creator's Script Builder -> Export, one
                         file with every form's field definitions AND every
                         workflow's push-to-API Deluge script).

0b) auto-map-ds      -- the MOST ACCURATE option. Feed it the .ds file, the
                         form name, and your raw .xlsx. It reads the form's
                         own field definitions straight out of the .ds export
                         (type=, values=, displayformat=) together with its
                         push workflow script, so it settles a Lookup field's
                         exact import format -- bare Code, bare Name, or a
                         Name-Code/Code-Name combo with a specific separator
                         -- with CERTAINTY instead of a fuzzy-matched guess.
                         It also auto-picks the live/production workflow when
                         a form has more than one (pass --workflow-var to
                         override), auto-detects Yes/No and short-code
                         picklists that need "map"/"yesno" conversion from
                         the workflow's .subString() calls, and clearly flags
                         any field wrapped in a custom thisapp.Dropdown_Fields
                         function as unverified (that function's source isn't
                         in the .ds export, so it can never be more than
                         "plain" + a warning). Writes a config that's usually
                         ready to run as-is -- scan it for "TODO" in
                         hedb_sheet/raw_column and any "UNVERIFIED
                         custom-function" note, fix just those, then `build`.

THE FAST PATH FOR A NEW FORM (script only, no .ds export)
------------------------------------------------------------
0) auto-map         -- feed it the form's script + your raw .xlsx + HEDB_.xlsx.
                        It fuzzy-matches every input.X field to a real column
                        in your raw file, and for lookup fields, tests your
                        actual sample values against every candidate HEDB
                        sheet (registry first, then fuzzy sheet-name + layout
                        auto-detection) to figure out, from the data itself,
                        whether conversion is even needed and which sheet to
                        use. Writes a config that's usually ready to run
                        as-is -- scan it for "TODO" / low-confidence
                        "_match_method" notes, fix just those, then `build`.
                        This replaces manually filling in analyze-script's
                        output for most fields; you still won't be able to
                        avoid confirming a Lookup field's Display Fields
                        setting in Zoho for anything auto-map flags as
                        low-confidence, or a custom-function field's source.
                        Less accurate than auto-map-ds above since it can
                        only see the workflow script, not the form's own
                        displayformat= configuration -- use auto-map-ds
                        instead whenever you have a .ds export available.

THE MANUAL 3-STEP WORKFLOW (fallback / for understanding what auto-map does)
------------------------------------------------------------------------------
1) analyze-script   -- feed it the form's Deluge add/update script (as a
                        .txt file). It extracts every `input.X` used in
                        the payload and classifies each as:
                          lookup  -> input.Field.Code               (needs HEDB Name text)
                          plain   -> input.Field                    (passthrough, just rename)
                          custom  -> thisapp.Fn(input.Field)         (needs that function's source -- flag for follow-up)
                        You'll also want to hand-mark any Date fields with
                        "kind": "date" (see Department_joining_Date in the
                        example config) -- Zoho's Date fields choke on a
                        date-TIME value like "2024-11-01 00:00:00"; this
                        kind strips the time and forces a clean yyyy-mm-dd
                        cell format on output.
                        It writes a starter mapping-config JSON you edit
                        by hand (fill in raw_column + hedb_sheet).

2) describe-sheet   -- if a lookup field's HEDB sheet isn't already in
                        the built-in registry (see LOOKUP_SHEET_REGISTRY
                        below), use this to dump the sheet's header row
                        and first few data rows so you can see which
                        column is the code and which is the name, then
                        add a short entry to the registry or override it
                        inline in the mapping config (code_col/name_col).

3) build            -- runs the actual transform: renames columns,
                        converts lookup codes -> HEDB Name text, leaves
                        plain fields untouched, and writes an
                        Import_Ready.xlsx with a trailing _NOTES column
                        flagging anything unresolved or still "pending".

USAGE
-----
python3 cheds_import_tool.py list-forms zoho_app.ds
python3 cheds_import_tool.py auto-map-ds zoho_app.ds Some_Form_Name \
    --raw raw_data.xlsx --hedb HEDB_.xlsx --out mapping.json
python3 cheds_import_tool.py analyze-script my_form_script.txt --out mapping.json
python3 cheds_import_tool.py describe-sheet HEDB_.xlsx "Some Sheet Name"
python3 cheds_import_tool.py build --raw raw_data.xlsx --config mapping.json \
    --hedb HEDB_.xlsx --out Import_Ready.xlsx

Requires: openpyxl (pip install openpyxl --break-system-packages)
"""

import sys
import re
import json
import argparse
import datetime
import difflib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# .ds (whole-application Deluge/script export) extraction.
# ---------------------------------------------------------------------------
# Zoho Creator can export an entire application as one big ".ds" script file
# containing every form's field definitions (type, values, displayformat)
# AND every workflow's Deluge script (the invokeurl push-to-CHEDS logic), all
# in one place. That's strictly better than pasting the form script and the
# workflow script separately for each form -- displayformat tells us the
# exact Lookup combo order/separator with certainty instead of guessing from
# screenshots, and it covers all forms in the app at once.
def _extract_brace_block(text, open_brace_idx):
    """Given the index of an opening '{', return the substring up to and
    including its matching closing '}', by tracking brace depth."""
    depth = 0
    for i in range(open_brace_idx, len(text)):
        if text[i] == '{':
            depth += 1
        elif text[i] == '}':
            depth -= 1
            if depth == 0:
                return text[open_brace_idx:i + 1]
    return None


def _extract_paren_block(text, open_paren_idx):
    """Same idea as _extract_brace_block but for a field's '( ... )' property
    block inside a form definition."""
    depth = 0
    for i in range(open_paren_idx, len(text)):
        if text[i] == '(':
            depth += 1
        elif text[i] == ')':
            depth -= 1
            if depth == 0:
                return text[open_paren_idx:i + 1]
    return None


def list_ds_forms(ds_text):
    return re.findall(r'\bform\s+(\w+)\s*\n\s*\{', ds_text)


def extract_ds_form_block(ds_text, form_name):
    """The form's own field-definition block (type=picklist, values=, etc)."""
    m = re.search(rf'\bform\s+{re.escape(form_name)}\s*\n\s*\{{', ds_text)
    if not m:
        return None
    brace_idx = ds_text.index('{', m.end() - 1)
    return _extract_brace_block(ds_text, brace_idx)


def extract_ds_workflow_blocks(ds_text, form_name):
    """Every named workflow block ('push_apiNN as "..." { ... }') whose body
    references this form via 'form = <form_name>' -- this is where the
    invokeurl/API-push Deluge script lives, separate from the form's own
    field-definition block above."""
    results = []
    for m in re.finditer(r'(\w+)\s+as\s+"([^"]*)"\s*\n\s*\{', ds_text):
        brace_idx = ds_text.index('{', m.end() - 1)
        block = _extract_brace_block(ds_text, brace_idx)
        if block and re.search(rf'form\s*=\s*{re.escape(form_name)}\b', block):
            results.append((m.group(1), m.group(2), block))
    return results


def parse_displayformat(expr):
    """Parse the contents of a field's `displayformat = [ ... ]` expression
    (the single most reliable signal for how a Lookup field must be
    imported -- it's exactly what Zoho renders in the picklist, so it's
    better evidence than guessing from a screenshot or from raw sample data).
    Returns one of:
      {"kind": "code"}                                    -- bare Code
      {"kind": "name"}                                     -- bare Name
      {"kind": "combo", "combo_order", "separator", "leading"}
      {"kind": "other", "raw": expr}                       -- neither Code nor
                                                                Name alone (e.g.
                                                                a custom column
                                                                like CIP_Family_Name)
    """
    e = expr.strip()
    if e == "Code":
        return {"kind": "code"}
    if e == "Name":
        return {"kind": "name"}
    if "+" in e and re.search(r"\bCode\b", e) and re.search(r"\bName\b", e):
        code_pos = re.search(r"\bCode\b", e).start()
        name_pos = re.search(r"\bName\b", e).start()
        combo_order = "code_name" if code_pos < name_pos else "name_code"
        literals = re.findall(r'"([^"]*)"', e)
        separator = literals[-1] if literals else " - "
        leading = literals[:-1]
        return {"kind": "combo", "combo_order": combo_order, "separator": separator,
                "leading": leading, "raw": e}
    return {"kind": "other", "raw": e}


# Field/section names that are structural, never real data fields, even
# though they appear as blocks in the .form export.
_FORM_STRUCTURAL_NAMES = {"section", "system_data"}


def parse_form_fields(form_text):
    """
    Parse a form's field-definition block (as written by `extract-form`, or
    the raw block from extract_ds_form_block) into
    {field_name: {"type", "values_expr", "report_name", "displayformat",
                  "literal_choices"}}.
    This is what lets auto-map-ds know a Lookup field's exact Display Fields
    configuration (bare Code / bare Name / a Name-Code combo with a specific
    separator) straight from the .ds export, instead of guessing.
    """
    # Stop before the "actions { on add { submit ( ... ) } }" tail -- those
    # parens are button definitions, not data fields.
    cut = re.search(r'\n\s*actions\s*\n\s*\{', form_text)
    body = form_text[:cut.start()] if cut else form_text

    fields = {}
    for m in re.finditer(r'(?:must have\s+)?([A-Za-z_][A-Za-z0-9_]*)\s*\n\s*\(', body):
        name = m.group(1)
        if name.lower() in _FORM_STRUCTURAL_NAMES:
            continue
        open_idx = body.index('(', m.end() - 1)
        block = _extract_paren_block(body, open_idx)
        if block is None:
            continue

        type_m = re.search(r'\btype\s*=\s*(\w+)', block)
        ftype = type_m.group(1).lower() if type_m else None
        if ftype == "section":
            continue

        values_m = re.search(r'\bvalues\s*=\s*(.+)', block)
        values_expr = values_m.group(1).strip() if values_m else None

        report_name = None
        literal_choices = None
        if values_expr:
            if values_expr.startswith("{"):
                literal_choices = re.findall(r'"([^"]*)"', values_expr)
            else:
                rm = re.match(r'([A-Za-z_][A-Za-z0-9_]*)', values_expr)
                if rm:
                    report_name = rm.group(1)

        df_m = re.search(r'\bdisplayformat\s*=\s*\[(.*?)\]', block)
        displayformat = parse_displayformat(df_m.group(1)) if df_m else None

        # A field can appear twice in a .form export only if there's a real
        # link-name collision (e.g. Skills_Description1 = a system field
        # sharing a display label) -- last one wins here, same as Zoho would
        # resolve input.X to the actual field with that link name.
        fields[name] = {
            "type": ftype,
            "values_expr": values_expr,
            "report_name": report_name,
            "displayformat": displayformat,
            "literal_choices": literal_choices,
        }
    return fields


def cmd_list_forms(args):
    with open(args.ds, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()
    for name in list_ds_forms(text):
        print(name)


def cmd_extract_form(args):
    with open(args.ds, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()

    form_block = extract_ds_form_block(text, args.form)
    if form_block is None:
        forms = list_ds_forms(text)
        close = difflib.get_close_matches(args.form, forms, n=5, cutoff=0.3)
        print(f"Form '{args.form}' not found.", file=sys.stderr)
        if close:
            print("Did you mean one of:", close, file=sys.stderr)
        sys.exit(1)

    workflows = extract_ds_workflow_blocks(text, args.form)

    prefix = args.out_prefix or args.form
    fields_path = f"{prefix}_fields.form"
    with open(fields_path, "w", encoding="utf-8") as f:
        f.write(f"form {args.form}\n" + form_block)
    print(f"Wrote field definitions to {fields_path}")

    if not workflows:
        print("No push-to-API workflow block found referencing this form "
              "(it may not have one, or may be named differently) -- "
              "check manually if you expected one.", file=sys.stderr)
    for i, (var_name, label, block) in enumerate(workflows):
        suffix = "" if len(workflows) == 1 else f"_{i+1}"
        wf_path = f"{prefix}_workflow{suffix}.txt"
        with open(wf_path, "w", encoding="utf-8") as f:
            f.write(f"{var_name} as \"{label}\"\n" + block)
        print(f"Wrote workflow ({var_name} / \"{label}\") to {wf_path}")

    print(
        "\nNext: pull the Deluge script out of the workflow file's "
        "'custom deluge script ( ... )' block (or just point analyze-script/"
        "auto-map at the whole workflow file -- input.X pattern matching "
        "still works fine even with the surrounding form=/type= wrapper), "
        "and use the _fields.form file's values=/displayformat= lines to "
        "settle lookup_direction/combo_order/combo_separator/value_map for "
        "each field with certainty."
    )


# ---------------------------------------------------------------------------
# Registry of HEDB "List of Values" sheets we've already reverse-engineered.
# Each entry: (header_row, code_col, name_col). Column numbers are 1-based,
# counting the sheet's real columns (these HEDB sheets typically have a
# blank column A, so code/name usually start at column 2/3).
# Add to this as you map more sheets with `describe-sheet`.
# ---------------------------------------------------------------------------
LOOKUP_SHEET_REGISTRY = {
    "Institutional Codes":        {"header_row": 4, "code_col": 2, "name_col": 3},
    "Academic Period":            {"header_row": 4, "code_col": 2, "name_col": 3},
    "Area of Specialization":     {"header_row": 4, "code_col": 2, "name_col": 3},
    "Degree or Program Level":    {"header_row": 4, "code_col": 2, "name_col": 3},
    "Full Part Time":             {"header_row": 4, "code_col": 2, "name_col": 3},
    "Faculty Category":           {"header_row": 4, "code_col": 2, "name_col": 3},
    "Staff Position":             {"header_row": 4, "code_col": 2, "name_col": 3},
    # Country sheet is reversed: Nationality (name) is col 2, Code is col 3.
    "Country":                    {"header_row": 4, "code_col": 3, "name_col": 2},
    # Universities sheet has no leading blank column.
    "Universities":               {"header_row": 4, "code_col": 1, "name_col": 2},
    # Fields of R&D: header on row 2 (not 4), and it's a two-level hierarchy
    # (Domain rows use col2/col3, sub-Field rows use col1/col4/col5) crammed
    # into one sheet. Using col1(code)/col5(name) naturally skips the Domain
    # summary rows (their col5 is blank) and keeps only the real sub-Field
    # entries, which is what a "Research_Area" field actually needs.
    "Fields of R&D":              {"header_row": 2, "code_col": 1, "name_col": 5},
    # Program Master: columns are Institution code / Institution Name /
    # CAA Program Code / Old Program Code / Program Name / Concentration /
    # Program Status. The program identifier that raw exports actually use
    # (format "141-6511-999999999") is the CAA Program Code in column 3, and
    # the descriptive text a Lookup field's "Name" wants is column 5 -- NOT
    # columns 2/3 like the generic default assumes. This is also very likely
    # the correct sheet for PLO's Program_Code field (same code format seen
    # there), superseding the earlier guess of "Program Level".
    "Program Master":             {"header_row": 4, "code_col": 3, "name_col": 5},
    # Student Research Type: no "Back to Main Menu" spacer column here --
    # header is row 2, Code in col 1, Description in col 2 (and unusually,
    # the Description text itself repeats the code as a prefix, e.g.
    # "FL  Faculty-led research projects" -- that's correct, not a bug).
    "Student Research Type":      {"header_row": 2, "code_col": 1, "name_col": 2},
}


# Maps the Zoho "report" name used in a Lookup field's `values = Report.ID`
# expression (e.g. "Institutional_Codes", "Program_Code") to the matching
# HEDB_.xlsx sheet name in LOOKUP_SHEET_REGISTRY above. This is what lets
# auto-map-ds resolve a lookup field's hedb_sheet straight from the .ds
# export's field definitions, without needing to fuzzy-match sample data.
# Add an entry here whenever you confirm a new report <-> HEDB sheet pairing.
REPORT_TO_HEDB_SHEET = {
    "institutionalcodes": "Institutional Codes",
    "academicperiod": "Academic Period",
    "countries": "Country",
    "degreeprogramlevel": "Degree or Program Level",
    "programcode": "Program Master",
    "areaofspecialization": "Area of Specialization",
    "universities": "Universities",
    "studentresearchtype": "Student Research Type",
    "fieldsofrd": "Fields of R&D",
    # Seen in the .ds export but not yet reverse-engineered against a HEDB
    # sheet -- confirm the matching "List of Values" tab with `describe-sheet`
    # and add a real entry above once you have it.
    "academicyear": None,
    "academicproficiencytests": None,
}


def read_text_robust(path):
    """
    Read a text file trying several encodings. Windows PowerShell's `>`
    redirection (and Notepad's default 'Save As') often writes UTF-16 or
    UTF-8-with-BOM instead of plain UTF-8, which trips up a plain
    open(path).read(). Try the common ones in order before giving up.
    """
    for enc in ("utf-8-sig", "utf-16", "utf-8", "cp1252"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    # last resort: ignore bad bytes rather than crash
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def norm(v):
    """Normalize a cell value for matching: '141.0' / 141.0 / '141' all -> '141'."""
    if v is None:
        return ""
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
        return str(f)
    except (TypeError, ValueError):
        return str(v).strip()


# ---------------------------------------------------------------------------
# Step 2: describe-sheet
# ---------------------------------------------------------------------------
def cmd_describe_sheet(args):
    wb = openpyxl.load_workbook(args.hedb, data_only=True)
    if args.sheet not in wb.sheetnames:
        matches = [s for s in wb.sheetnames if args.sheet.lower() in s.lower()]
        print(f"Sheet '{args.sheet}' not found.", file=sys.stderr)
        if matches:
            print("Did you mean:", matches, file=sys.stderr)
        sys.exit(1)
    ws = wb[args.sheet]
    print(f"Sheet: {args.sheet}  (dims={ws.dimensions})")
    for r in range(1, min(args.rows, ws.max_row) + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, min(12, ws.max_column) + 1)]
        if any(v is not None for v in row):
            print(f"row {r}: {row}")
    print(
        "\nLook for the row with column headers like 'Code' / 'Name' / "
        "'Description' above, then note which column number holds the code "
        "and which holds the descriptive name. Add those as header_row/"
        "code_col/name_col in LOOKUP_SHEET_REGISTRY, or as a per-field "
        "override in your mapping config."
    )


# ---------------------------------------------------------------------------
# Step 1: analyze-script
# ---------------------------------------------------------------------------
def strip_deluge_comments(text):
    """Remove '// ...' line comments before parsing, so commented-out (dead)
    code doesn't get picked up as a live field reference. The negative
    lookbehind for ':' avoids treating "https://..." inside invokeurl string
    literals as a comment marker."""
    return re.sub(r"(?<!:)//[^\n]*", "", text)


# System/response fields that are ALWAYS blank on a bulk "add" import (they
# only get written back by the workflow *after* a successful API call, e.g.
# recordId, lastUpdated, or the raw API response/error/payload dumps). Every
# form spells these slightly differently (Record_id vs Record_ID, lastUpdated
# vs Last_Updated, etc.) so this is compared against the fully normalized
# (lowercase, alnum-only) field name everywhere it's used, never a literal
# string match.
SKIP_ALWAYS_KEYS = {
    "recordid", "payload", "apiresponse", "response", "error",
    "lastupdated", "lastupdateddatetime", "createdtime", "modifiedtime",
}


def classify_script_fields(text):
    """Parse a Deluge script's text and bucket every input.X reference into
    lookup / custom(pending) / plain, purely from usage pattern."""
    text = strip_deluge_comments(text)
    skip_always = SKIP_ALWAYS_KEYS

    # input.Field.Code           -> lookup field
    lookup_fields = sorted(f for f in set(re.findall(r"input\.([A-Za-z0-9_]+)\.Code\b", text))
                           if _norm_key(f) not in skip_always)
    # thisapp.Something(input.Field [, ...])  -> custom-converted field
    custom_calls = re.findall(r"thisapp\.[\w.]+\(([^()]*)\)", text)
    custom_fields = sorted(set(
        m.group(1) for call in custom_calls
        for m in re.finditer(r"input\.([A-Za-z0-9_]+)", call)
        if _norm_key(m.group(1)) not in skip_always
    ))
    # every other bare input.Field reference, excluding system/response
    # fields that are ALWAYS blank on a bulk "add" import (they only get
    # written back by the workflow *after* a successful API call, e.g.
    # recordId, lastUpdated, or the raw API response/error/payload dumps).
    # Every form spells these slightly differently -- Record_id vs
    # Record_ID, lastUpdated vs Last_Updated, etc. -- so compare on the
    # fully normalized (lowercase, alnum-only) form, not a literal string,
    # otherwise e.g. "Last_Updated" slips through as a "plain" field and
    # gets a nonsense best-guess raw-column suggestion.
    all_fields = sorted(set(re.findall(r"input\.([A-Za-z0-9_]+)", text)))
    plain_fields = [f for f in all_fields
                    if f not in lookup_fields and f not in custom_fields
                    and _norm_key(f) not in skip_always]
    return lookup_fields, custom_fields, plain_fields


def cmd_analyze_script(args):
    text = read_text_robust(args.script)
    lookup_fields, custom_fields, plain_fields = classify_script_fields(text)

    fields = []
    for f in lookup_fields:
        fields.append({
            "raw_column": "TODO_raw_header_for_" + f,
            "target_field": f,
            "kind": "lookup",
            "hedb_sheet": "TODO -- pick matching sheet name from HEDB_.xlsx "
                          "(use `describe-sheet` if unsure)",
        })
    for f in custom_fields:
        fields.append({
            "raw_column": "TODO_raw_header_for_" + f,
            "target_field": f,
            "kind": "pending",
            "note": "value is passed through a custom Deluge function before "
                    "the API call -- get that function's source before trusting "
                    "raw codes here",
        })
    for f in plain_fields:
        fields.append({
            "raw_column": "TODO_raw_header_for_" + f,
            "target_field": f,
            "kind": "plain",
        })

    payload = json.dumps({"fields": fields}, indent=2)
    summary = (
        f"Found {len(lookup_fields)} lookup field(s), {len(custom_fields)} "
        f"custom-function field(s), {len(plain_fields)} plain field(s)."
    )

    if args.out:
        # Write with explicit UTF-8 -- avoids the classic Windows/PowerShell
        # issue where `> file.json` redirection saves as UTF-16 with a BOM,
        # which json.load() then can't parse.
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(payload)
        print(summary)
        print(f"Wrote starter config to {args.out}")
        print("Edit it: fill in each raw_column with the actual header text "
              "in your raw export, and hedb_sheet for each lookup field.")
    else:
        print(payload)
        print(
            f"\n// {summary}\n"
            "// Edit this JSON: fill in each raw_column with the actual header "
            "text in your raw export, and hedb_sheet for each lookup field.",
            file=sys.stderr,
        )


# ---------------------------------------------------------------------------
# Step 2 (automated): auto-map
# ---------------------------------------------------------------------------
def _norm_key(s):
    """lowercase, alnum-only -- for comparing 'Load_Academic_Period' to 'Academic_Period'."""
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def _pair_score(target_field, header):
    tnorm, hnorm = _norm_key(target_field), _norm_key(header)
    if not hnorm:
        return 0.0
    if tnorm == hnorm:
        return 1.0
    score = difflib.SequenceMatcher(None, tnorm, hnorm).ratio()
    # containment bonus: 'Academic_Period' inside 'Load_Academic_Period' --
    # but ONLY reward it, don't let it override a case where the raw header
    # is actually a better textual match for a *different* target (that's
    # handled by the global assignment below, not by inflating this score
    # past what plain similarity already says).
    if tnorm in hnorm or hnorm in tnorm:
        score = max(score, 0.85)
    return score


# We deliberately auto-accept only very confident matches. Testing this
# against real HEDB forms showed that a handful of fields have raw headers
# and Zoho field names that share NO real vocabulary at all (e.g. Zoho field
# "Admin_Workload1" vs raw header "Load_Career_Related_Credits" -- correct
# pairing, but a human wrote two totally different words for the same
# concept). No string-similarity score can bridge that safely, and guessing
# wrong on government compliance data is worse than asking a human. So:
# anything below this bar is left for manual review rather than silently
# auto-assigned, even though a *lower*-confidence guess is shown as a hint.
AUTO_ACCEPT_SCORE = 0.80


def match_all_headers(target_fields, raw_headers, accept_threshold=AUTO_ACCEPT_SCORE):
    """
    Global (not per-field-independent) matching: score every
    (target_field, raw_header) pair, then greedily assign the highest-scoring
    pairs first, retiring both sides once used. This is what prevents two
    different target fields (e.g. Maximum_Workload_per_semester1 and
    Admin_Workload1) from both independently claiming the same raw column.
    Returns:
      accepted:   {target_field: (header, score)}       -- score >= threshold
      best_guess: {target_field: (header, score)}       -- best available,
                  even if below threshold, for hinting a human reviewer
    """
    real_headers = [h for h in raw_headers if h]
    pairs = sorted(
        ((_pair_score(t, h), t, h) for t in target_fields for h in real_headers),
        key=lambda x: -x[0],
    )
    accepted, best_guess, used_headers = {}, {}, set()
    for score, t, h in pairs:
        if t not in best_guess:
            best_guess[t] = (h, score)
        if t in accepted or h in used_headers:
            continue
        if score < accept_threshold:
            continue
        accepted[t] = (h, score)
        used_headers.add(h)
    return accepted, best_guess


def detect_sheet_layout(ws, max_rows=10, max_cols=15):
    """Generic fallback for HEDB sheets not already in LOOKUP_SHEET_REGISTRY:
    scan the first few rows for a header containing 'code', then pick the
    nearest other text header on that row as the name/description column."""
    for r in range(1, min(max_rows, ws.max_row) + 1):
        texts = []
        for c in range(1, min(max_cols, ws.max_column) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                texts.append((c, v.strip().lower()))
        code_col = next((c for c, v in texts if "code" in v), None)
        if code_col is None:
            continue
        others = [c for c, v in texts if c != code_col]
        if not others:
            continue
        name_col = min(others, key=lambda c: abs(c - code_col))
        return r, code_col, name_col
    return None


def _overlap_score(sample_values, code_name_map):
    """What fraction of sample_values are found as CODE keys vs as NAME
    values in a code->name map? Returns (code_overlap, name_overlap)."""
    if not sample_values:
        return 0.0, 0.0
    keys = code_name_map.keys()
    vals_lower = {v.lower() for v in code_name_map.values()}
    code_hits = sum(1 for s in sample_values if norm(s) in keys)
    name_hits = sum(1 for s in sample_values if str(s).strip().lower() in vals_lower)
    n = len(sample_values)
    return code_hits / n, name_hits / n


def resolve_lookup_field(sample_values, hedb_wb, target_field, min_overlap=0.6):
    """
    Figure out, purely from the data, how a lookup field should be handled:
      - if the raw sample values already match a HEDB sheet's descriptive
        Name column, no conversion is needed (kind should be "plain").
      - if they match a sheet's Code column instead, mark as "lookup" so
        `build` converts code -> name.
      - otherwise, fall back to fuzzy sheet-name matching and flag as
        low-confidence for manual review.
    Returns a dict describing the best guess, or None.
    """
    sample_values = [s for s in sample_values if s not in (None, "")]
    if not sample_values:
        return None

    candidates = []

    # 1) trust the hand-verified registry first
    for sheet, cfg in LOOKUP_SHEET_REGISTRY.items():
        if sheet not in hedb_wb.sheetnames:
            continue
        m = load_lookup_map(hedb_wb, sheet, cfg.get("header_row"), cfg.get("code_col"), cfg.get("name_col"))
        code_ov, name_ov = _overlap_score(sample_values, m)
        candidates.append((max(code_ov, name_ov), code_ov, name_ov, sheet, cfg, "registry"))

    # 2) fuzzy-match sheet names against the field name, try a handful of the
    #    closest ones with generic layout detection (skip if registry already
    #    found something solid, to keep this fast)
    if not candidates or max(c[0] for c in candidates) < min_overlap:
        name_guesses = difflib.get_close_matches(
            target_field.replace("_", " "), hedb_wb.sheetnames, n=5, cutoff=0.3
        )
        for sheet in name_guesses:
            if sheet in LOOKUP_SHEET_REGISTRY:
                continue
            layout = detect_sheet_layout(hedb_wb[sheet])
            if not layout:
                continue
            header_row, code_col, name_col = layout
            cfg = {"header_row": header_row, "code_col": code_col, "name_col": name_col}
            m = load_lookup_map(hedb_wb, sheet, header_row, code_col, name_col)
            code_ov, name_ov = _overlap_score(sample_values, m)
            candidates.append((max(code_ov, name_ov), code_ov, name_ov, sheet, cfg, "fuzzy-name+auto-layout"))

    if not candidates:
        return None
    candidates.sort(key=lambda c: c[0], reverse=True)
    best_score, code_ov, name_ov, sheet, cfg, method = candidates[0]
    if best_score < min_overlap:
        return {
            "confident": False,
            "suggestions": [c[3] for c in candidates[:3]],
            "best_score": round(best_score, 2),
        }
    return {
        "confident": True,
        "hedb_sheet": sheet,
        "cfg": cfg,
        "match_type": "name" if name_ov >= code_ov else "code",
        "score": round(best_score, 2),
        "method": method,
    }


def cmd_auto_map(args):
    text = read_text_robust(args.script)
    lookup_fields, custom_fields, plain_fields = classify_script_fields(text)
    all_target_fields = lookup_fields + custom_fields + plain_fields

    raw_wb = openpyxl.load_workbook(args.raw, data_only=True)
    raw_ws = raw_wb[raw_wb.sheetnames[0]] if not args.sheet else raw_wb[args.sheet]
    raw_headers = [raw_ws.cell(row=1, column=c).value for c in range(1, raw_ws.max_column + 1)]

    # ONE global assignment across every field on the form at once -- this is
    # what stops e.g. "Admin_Workload1" and "Maximum_Workload_per_semester1"
    # (or any lookup/plain/custom field) from independently grabbing the same
    # raw column just because they happen to share a word like "workload".
    accepted, best_guess = match_all_headers(all_target_fields, raw_headers)

    def sample_column(header, n=25):
        if header is None:
            return []
        col = raw_headers.index(header) + 1
        out = []
        for r in range(2, raw_ws.max_row + 1):
            v = raw_ws.cell(row=r, column=col).value
            if v not in (None, ""):
                out.append(v)
            if len(out) >= n:
                break
        return out

    hedb_wb = openpyxl.load_workbook(args.hedb, data_only=True) if args.hedb else None

    fields = []
    counts = {"auto_lookup": 0, "auto_plain_already_named": 0, "needs_review": 0, "pending": 0, "plain": 0}

    def resolve_raw_column(f):
        """Returns (header_to_use, method_note, confident: bool)."""
        if f in accepted:
            h, s = accepted[f]
            return h, f"raw_column auto-matched (score {s:.2f})", True
        guess, gscore = best_guess.get(f, (None, 0.0))
        if guess:
            return (
                f"TODO -- verify then set to '{guess}' (best guess, score {gscore:.2f}, "
                f"below the {AUTO_ACCEPT_SCORE} auto-accept bar -- these two names don't "
                f"share enough text for the tool to trust it automatically)",
                f"low-confidence guess only ({gscore:.2f})",
                False,
            )
        return "TODO_raw_header_for_" + f, "no candidate raw column found at all", False

    for f in lookup_fields:
        header, method, confident = resolve_raw_column(f)
        if not confident or hedb_wb is None:
            fields.append({
                "raw_column": header,
                "target_field": f,
                "kind": "lookup",
                "hedb_sheet": "TODO -- pick matching sheet name from HEDB_.xlsx",
                "_match_method": method if confident else method + " -- fill in raw_column first",
            })
            counts["needs_review"] += 1
            continue
        samples = sample_column(header)
        resolution = resolve_lookup_field(samples, hedb_wb, f)
        if resolution and resolution.get("confident"):
            if resolution["match_type"] == "name":
                fields.append({
                    "raw_column": header,
                    "target_field": f,
                    "kind": "plain",
                    "_match_method": f"raw values already match '{resolution['hedb_sheet']}' Name column "
                                     f"({resolution['method']}, overlap {resolution['score']}) -- no conversion needed",
                })
                counts["auto_plain_already_named"] += 1
            else:
                fields.append({
                    "raw_column": header,
                    "target_field": f,
                    "kind": "lookup",
                    "hedb_sheet": resolution["hedb_sheet"],
                    "_match_method": f"auto-detected via {resolution['method']}, code-overlap {resolution['score']}",
                })
                counts["auto_lookup"] += 1
        else:
            suggestions = resolution["suggestions"] if resolution else []
            fields.append({
                "raw_column": header,
                "target_field": f,
                "kind": "lookup",
                "hedb_sheet": "TODO -- low confidence, candidates: " + (", ".join(suggestions) or "none found"),
                "_match_method": f"low overlap ({resolution['best_score'] if resolution else 0}) -- verify manually, "
                                  "or run `describe-sheet` on one of the candidates",
            })
            counts["needs_review"] += 1

    for f in custom_fields:
        header, method, confident = resolve_raw_column(f)
        fields.append({
            "raw_column": header,
            "target_field": f,
            "kind": "pending",
            "note": "value is passed through a custom Deluge function before the API call -- "
                    "get that function's source before trusting raw codes here",
            "_match_method": method,
        })
        counts["pending"] += 1
        if not confident:
            counts["needs_review"] += 1

    for f in plain_fields:
        header, method, confident = resolve_raw_column(f)
        entry = {"raw_column": header, "target_field": f, "kind": "plain"}
        if not confident:
            entry["_match_method"] = method
            counts["needs_review"] += 1
        counts["plain"] += 1
        fields.append(entry)

    payload = json.dumps({"fields": fields}, indent=2)
    summary = (
        f"Auto-resolved: {counts['auto_lookup']} lookup field(s) needing conversion, "
        f"{counts['auto_plain_already_named']} lookup field(s) whose raw data was already correct text, "
        f"{counts['plain']} plain field(s), {counts['pending']} custom-function field(s) (always need manual review). "
        f"{counts['needs_review']} entr(y/ies) still need a human look (see _match_method / TODO markers)."
    )

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(payload)
        print(summary)
        print(f"Wrote auto-mapped config to {args.out}")
        print("Scan for 'TODO' and low-confidence '_match_method' notes before running `build`; "
              "everything else is ready to go as-is.")
    else:
        print(payload)
        print(f"\n// {summary}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Step 2 (most accurate): auto-map-ds
# ---------------------------------------------------------------------------
# auto-map builds a config purely from a workflow script's input.X.Code /
# thisapp.Fn() / bare input.X usage patterns, then leans on fuzzy sample-data
# matching to guess a Lookup field's HEDB sheet. That's a reasonable fallback,
# but the .ds export actually TELLS us the answer for most of that guesswork:
# every Lookup field's `displayformat = [...]` line is exactly what Zoho
# renders in the picklist (bare Code, bare Name, or a Name-Code/Code-Name
# combo with a specific separator), and its `values = Report.ID` line names
# the exact linked report. auto-map-ds reads BOTH the form's field-definition
# block and its push workflow straight out of one whole-app .ds export, so it
# can settle lookup_direction/combo_order/combo_separator/hedb_sheet with
# certainty instead of a confidence score -- the same way every form in this
# app was hand-verified earlier. It still can't see inside a custom
# thisapp.Dropdown_Fields.X() function's source, so those fields are always
# flagged "plain" + a note rather than guessed.
def _pick_ds_workflow(workflows, requested_var=None):
    """workflows: list of (var_name, label, block) from extract_ds_workflow_blocks.
    Picks the one to treat as authoritative: an explicit --workflow-var match,
    else whichever's var_name/label mentions 'production', else the last one
    found (every form worked through by hand so far had its production/active
    workflow appear last in the export)."""
    if not workflows:
        return None, []
    if requested_var:
        for var_name, label, block in workflows:
            if var_name == requested_var:
                return (var_name, label, block), [w[0] for w in workflows]
        raise ValueError(f"--workflow-var '{requested_var}' not found among: "
                          f"{[w[0] for w in workflows]}")
    prod = [w for w in workflows if "production" in w[0].lower() or "production" in w[1].lower()]
    chosen = prod[-1] if prod else workflows[-1]
    return chosen, [w[0] for w in workflows]


def build_ds_field_entries(form_fields, workflow_text):
    """Combine a form's parsed field definitions with its push-workflow script
    to produce mapping-config field entries with high-confidence kind/
    hedb_sheet/lookup_direction -- the .ds-informed equivalent of auto-map's
    per-field guessing."""
    stripped = strip_deluge_comments(workflow_text)
    lookup_fields, custom_fields, plain_fields = classify_script_fields(stripped)
    lookup_set, custom_set, plain_set = set(lookup_fields), set(custom_fields), set(plain_fields)

    entries = []
    for name, fdef in form_fields.items():
        if _norm_key(name) in SKIP_ALWAYS_KEYS:
            continue

        ftype = fdef["type"]
        report_name = fdef["report_name"]
        df = fdef["displayformat"]
        choices = fdef["literal_choices"]
        referenced = name in lookup_set or name in custom_set or name in plain_set
        entry = {"raw_column": None, "target_field": name}
        notes = []

        substring_m = re.search(
            rf"input\.{re.escape(name)}\.(?:subString|SubString)\(\s*0\s*,\s*(\d+)\s*\)", stripped
        )

        if name in custom_set:
            entry["kind"] = "plain"
            notes.append(
                "UNVERIFIED custom-function field -- value is passed through a "
                "thisapp.Dropdown_Fields...() (or similar) custom function before "
                "the API call; that function's source isn't in the .ds export. "
                "Kept 'plain' -- correct only if raw data already holds the exact "
                "literal picklist choice text" + (f" ({choices})" if choices else ".")
            )
        elif substring_m and choices:
            n = int(substring_m.group(1))
            value_map = {c[:n]: c for c in choices}
            if n == 1 and {c.lower() for c in choices} == {"yes", "no"}:
                entry["kind"] = "yesno"
                notes.append(
                    f"Picklist choices = {choices}; workflow does .subString(0,{n}) "
                    "for the API. Using 'yesno' for flexible Y/N/true/false/1/0 raw input."
                )
            else:
                entry["kind"] = "map"
                entry["value_map"] = value_map
                notes.append(
                    f"Picklist choices = {choices}; workflow does .subString(0,{n}) "
                    f"to get a {n}-char code for the API. value_map assumes raw data "
                    "holds that short code -- if raw already has the full literal "
                    "text, switch this field to kind:'plain' instead."
                )
        elif name in lookup_set or (report_name and df):
            if df is None:
                entry["kind"] = "lookup"
                entry["hedb_sheet"] = "TODO -- no displayformat found in .ds export, verify Display Fields setting manually"
            elif df["kind"] == "code":
                entry["kind"] = "plain"
                notes.append("displayformat = [Code] confirmed in .ds export -- bare code, no conversion needed.")
            elif df["kind"] == "name":
                sheet = REPORT_TO_HEDB_SHEET.get(_norm_key(report_name)) if report_name else None
                entry["kind"] = "lookup"
                entry["lookup_direction"] = "code_to_name"
                entry["hedb_sheet"] = sheet or f"TODO -- report '{report_name}' not yet in REPORT_TO_HEDB_SHEET, find/register its HEDB sheet"
                notes.append("displayformat = [Name] confirmed in .ds export.")
            elif df["kind"] == "combo":
                sheet = REPORT_TO_HEDB_SHEET.get(_norm_key(report_name)) if report_name else None
                entry["kind"] = "lookup"
                entry["lookup_direction"] = "code_to_combo"
                entry["combo_order"] = df["combo_order"]
                entry["combo_separator"] = df["separator"]
                entry["hedb_sheet"] = sheet or f"TODO -- report '{report_name}' not yet in REPORT_TO_HEDB_SHEET, find/register its HEDB sheet"
                note = f"displayformat = [{df['raw']}] confirmed in .ds export -- combo order '{df['combo_order']}', separator {df['separator']!r}."
                if df["leading"]:
                    note += f" Extra literal text also present in the format string ({df['leading']!r}) -- double check it's cosmetic and not required in the imported value."
                notes.append(note)
            else:
                entry["kind"] = "lookup"
                entry["lookup_direction"] = "code_to_name"
                sheet = REPORT_TO_HEDB_SHEET.get(_norm_key(report_name)) if report_name else None
                entry["hedb_sheet"] = (sheet + " -- UNCONFIRMED, see note") if sheet else "TODO -- unusual displayformat, verify manually"
                notes.append(
                    f"displayformat = [{df['raw']}] doesn't reference bare Code or Name -- "
                    "it points at some other column on the linked report. Verify which HEDB "
                    "sheet/column that actually is before trusting this mapping."
                )
            if not referenced:
                notes.append(
                    "Not referenced anywhere in the push workflow script -- likely hardcoded "
                    "elsewhere (check for a literal value near institutionCode) or simply unused. "
                    "May still need a valid value to save the Zoho record."
                )
        elif ftype == "date":
            entry["kind"] = "date"
        elif ftype == "phonenumber":
            entry["kind"] = "phone"
            notes.append("type = phonenumber -- raw digits/dashes/spaces will be stripped to "
                         "a bare '+countrycode+number' string (e.g. '+971-50-1203650' -> "
                         "'+971501203650') to match Zoho's expected format.")
            if not referenced:
                notes.append("Not referenced in the push workflow script.")
        elif choices:
            entry["kind"] = "plain"
            notes.append(f"Literal picklist choices = {choices} -- raw data must match one exactly.")
            if not referenced:
                notes.append("Not referenced in the push workflow script.")
        else:
            entry["kind"] = "plain"
            if not referenced:
                notes.append("Not referenced in the push workflow script (may be unused/commented out there).")

        if notes:
            entry["_note"] = " ".join(notes)
        entries.append(entry)
    return entries


def cmd_auto_map_ds(args):
    ds_text = read_text_robust(args.ds)

    form_block = extract_ds_form_block(ds_text, args.form)
    if form_block is None:
        forms = list_ds_forms(ds_text)
        close = difflib.get_close_matches(args.form, forms, n=5, cutoff=0.3)
        print(f"Form '{args.form}' not found in {args.ds}.", file=sys.stderr)
        if close:
            print("Did you mean one of:", close, file=sys.stderr)
        sys.exit(1)

    workflows = extract_ds_workflow_blocks(ds_text, args.form)
    if not workflows:
        print(f"WARNING: no push-to-API workflow block found for form '{args.form}' -- "
              "this form may not have a live CHEDS integration yet. Building the config "
              "from field definitions only; every field will be marked as not referenced "
              "in any workflow.", file=sys.stderr)
        workflow_text = ""
        all_var_names = []
    else:
        (var_name, label, block), all_var_names = _pick_ds_workflow(workflows, args.workflow_var)
        print(f"Using workflow '{var_name}' (\"{label}\") out of {all_var_names} -- "
              f"pass --workflow-var to pick a different one.")
        workflow_text = block

    form_fields = parse_form_fields(form_block)

    raw_wb = openpyxl.load_workbook(args.raw, data_only=True)
    raw_ws = raw_wb[raw_wb.sheetnames[0]] if not args.sheet else raw_wb[args.sheet]
    raw_headers = [raw_ws.cell(row=1, column=c).value for c in range(1, raw_ws.max_column + 1)]

    entries = build_ds_field_entries(form_fields, workflow_text)

    target_fields = [e["target_field"] for e in entries]
    accepted, best_guess = match_all_headers(target_fields, raw_headers)
    needs_review = 0
    for e in entries:
        f = e["target_field"]
        if f in accepted:
            h, s = accepted[f]
            e["raw_column"] = h
        else:
            guess, gscore = best_guess.get(f, (None, 0.0))
            needs_review += 1
            if guess:
                e["raw_column"] = (
                    f"TODO -- verify then set to '{guess}' (best guess, score {gscore:.2f}, "
                    f"below the {AUTO_ACCEPT_SCORE} auto-accept bar)"
                )
            else:
                e["raw_column"] = "TODO_raw_header_for_" + f

    if args.hedb:
        hedb_wb = openpyxl.load_workbook(args.hedb, data_only=True)
        for e in entries:
            sheet = e.get("hedb_sheet")
            if sheet and not sheet.startswith("TODO") and sheet.split(" -- UNCONFIRMED")[0] not in hedb_wb.sheetnames:
                e["_note"] = (e.get("_note", "") + f" WARNING: sheet '{sheet}' not found in {args.hedb} -- check the exact sheet name.").strip()

    lookup_kinds = sum(1 for e in entries if e.get("kind") == "lookup")
    plain_kinds = sum(1 for e in entries if e.get("kind") == "plain")
    map_kinds = sum(1 for e in entries if e.get("kind") in ("map", "yesno"))
    date_kinds = sum(1 for e in entries if e.get("kind") == "date")

    payload = json.dumps({"fields": entries}, indent=2)
    summary = (
        f"Built {len(entries)} field(s) from the .ds export: {lookup_kinds} lookup, "
        f"{plain_kinds} plain, {map_kinds} map/yesno, {date_kinds} date. "
        f"{needs_review} field(s) need raw_column verified by hand (see TODO markers)."
    )

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(payload)
        print(summary)
        print(f"Wrote config to {args.out}")
        print("Scan for 'TODO' in hedb_sheet/raw_column and any '_note' mentioning "
              "'UNVERIFIED custom-function' before running `build` -- those still need "
              "a human look; everything else was settled directly from the .ds export's "
              "displayformat/values lines.")
    else:
        print(payload)
        print(f"\n// {summary}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Step 3: build
# ---------------------------------------------------------------------------
def load_lookup_map(hedb_wb, sheet_name, header_row=None, code_col=None, name_col=None):
    if sheet_name not in hedb_wb.sheetnames:
        raise ValueError(f"HEDB sheet '{sheet_name}' not found in workbook.")
    reg = LOOKUP_SHEET_REGISTRY.get(sheet_name, {})
    header_row = header_row or reg.get("header_row", 4)
    code_col = code_col or reg.get("code_col", 2)
    name_col = name_col or reg.get("name_col", 3)

    ws = hedb_wb[sheet_name]
    m = {}
    for r in range(header_row + 1, ws.max_row + 1):
        c = ws.cell(row=r, column=code_col).value
        n = ws.cell(row=r, column=name_col).value
        if c not in (None, "") and n not in (None, ""):
            m[norm(c)] = str(n).strip()
    return m


def build_reverse_map(code_to_name):
    """name(lowercased/stripped) -> (canonical_name, code). If two codes
    share a name (shouldn't happen for these HEDB lists, but just in case)
    the first one wins -- flagged separately if it ever matters."""
    rev = {}
    for code, name in code_to_name.items():
        key = name.strip().lower()
        if key not in rev:
            rev[key] = (name, code)
    return rev


# A Zoho Lookup field's dropdown often displays a combo of its Name and Code
# Display Fields -- but NOT always in the same order. Academic_Period showed
# "1990-1991 : Fall - 199001" (Name - Code); Student_Degree's Display Fields
# panel showed Code listed first, then Name ("Code - Name"). So the combo
# order has to be a setting too, not just whether to combine at all.
# `lookup_direction` controls which transform `build` applies to a
# "kind": "lookup" field:
#   code_to_name   (default) raw value is a CODE      -> output bare Name
#   name_to_code                raw value is a Name    -> output bare Code
#   code_to_combo                raw value is a CODE   -> output combo
#   name_to_combo                raw value is a Name    -> output combo
# `combo_order` ("name_code", the default, or "code_name") controls which
# half comes first in the combo; `combo_separator` (default " - ") controls
# the glue text -- match whatever Zoho's Display Fields panel shows for that
# specific field (order AND separator can differ field to field).
def apply_lookup_transform(v, code_to_name, name_to_code, direction, separator, combo_order="name_code"):
    def combo(name, code):
        return f"{code}{separator}{name}" if combo_order == "code_name" else f"{name}{separator}{code}"

    if direction == "name_to_code":
        hit = name_to_code.get(str(v).strip().lower())
        return (hit[1] if hit else None)
    if direction == "code_to_combo":
        name = code_to_name.get(norm(v))
        return (combo(name, norm(v)) if name else None)
    if direction == "name_to_combo":
        hit = name_to_code.get(str(v).strip().lower())
        return (combo(hit[0], hit[1]) if hit else None)
    # default: code_to_name
    return code_to_name.get(norm(v))


# Common across many HEDB forms: a raw export uses Y/N (or 1/0, True/False)
# for a field where Zoho's Radio/Dropdown choices are spelled out "Yes"/"No".
# "kind": "yesno" on a field handles this without needing any HEDB sheet at
# all -- pass "yes_label"/"no_label" in the field def if a field ever uses
# different exact wording than "Yes"/"No" (check the field's Choices list).
YES_VALUES = {"y", "yes", "true", "1"}
NO_VALUES = {"n", "no", "false", "0"}


def apply_yesno_transform(v, yes_label="Yes", no_label="No"):
    if v in (None, ""):
        return v, True  # nothing to convert, not an error
    key = str(v).strip().lower()
    if key in YES_VALUES:
        return yes_label, True
    if key in NO_VALUES:
        return no_label, True
    return v, False


def apply_phone_transform(v):
    """Zoho's phonenumber field type wants digits only (plus an optional
    leading +), e.g. '+971501203650' -- not '+971-50-1203650' or
    '+971 50 120 3650'. Strip every non-digit character except a leading '+'."""
    if v in (None, ""):
        return v
    s = str(v).strip()
    plus = s.startswith("+")
    digits = re.sub(r"\D", "", s)
    return ("+" if plus else "") + digits


def _lookup_cache_key(fdef):
    """
    Cache key for a field's lookup table. Deliberately more than just the
    sheet name: two different fields can legitimately point at the SAME
    hedb_sheet but different columns within it (e.g. 'Program Master' used
    once for its CIP_Family_Code column and again for its CIP Code column --
    found via a real bug where both fields silently shared one cached map,
    keyed only by sheet name, so the second field silently returned the
    first field's column). Keying on the full (sheet, header_row, code_col,
    name_col) tuple keeps each distinct lookup shape independent.
    """
    return (
        fdef.get("hedb_sheet", ""),
        fdef.get("header_row"),
        fdef.get("code_col"),
        fdef.get("name_col"),
    )


def preload_lookup_maps(config, hedb_wb):
    """
    Load a code->name map (and its reverse) for every distinct lookup
    "shape" (see _lookup_cache_key) a config's lookup fields reference.
    Shared by build_import_ready and build_review_report so both stay
    consistent. Don't let one unfinished/placeholder config entry (e.g. an
    un-edited "TODO -- pick matching sheet name..." left over from auto-map)
    blow up the whole run -- record the failure per-key and keep going.
    Returns (lookup_maps, reverse_maps, sheet_load_errors) -- all three
    dicts are keyed by _lookup_cache_key(fdef), NOT by sheet name alone.
    """
    lookup_maps = {}
    reverse_maps = {}
    sheet_load_errors = {}
    for fdef in config["fields"]:
        if fdef.get("kind") == "lookup":
            sheet_name = fdef.get("hedb_sheet", "")
            key = _lookup_cache_key(fdef)
            if key in lookup_maps or key in sheet_load_errors:
                continue
            try:
                m = load_lookup_map(
                    hedb_wb, sheet_name,
                    header_row=fdef.get("header_row"),
                    code_col=fdef.get("code_col"),
                    name_col=fdef.get("name_col"),
                )
                lookup_maps[key] = m
                reverse_maps[key] = build_reverse_map(m)
            except ValueError as e:
                sheet_load_errors[key] = str(e)
    return lookup_maps, reverse_maps, sheet_load_errors


def parse_date_value(v):
    """
    Shared date parsing used by both build_import_ready's "date" kind and
    build_review_report's cell classifier. Returns (ok, date_obj_or_None,
    had_time_component):
      ok=True,  had_time=True   -- parsed fine, but raw value carried a time
                                    portion that a Zoho Date field would choke
                                    on (e.g. "2024-11-01 00:00:00")
      ok=True,  had_time=False  -- parsed fine, already a clean date
      ok=False, ...             -- couldn't parse at all
    """
    if isinstance(v, datetime.datetime):
        return True, v.date(), True
    if isinstance(v, datetime.date):
        return True, v, False
    if isinstance(v, str) and v.strip():
        for fmt, had_time in (
            ("%Y-%m-%d %H:%M:%S", True),
            ("%Y-%m-%d", False),
            ("%m/%d/%Y", False),
            ("%d/%m/%Y", False),
        ):
            try:
                return True, datetime.datetime.strptime(v.strip(), fmt).date(), had_time
            except ValueError:
                continue
        return False, None, False
    return True, None, False  # blank -- nothing to parse, not an error


def build_import_ready(raw_path, config_path, hedb_path, out_path, sheet=None):
    """
    The actual raw-Excel -> import-ready-Excel transform, factored out of
    cmd_build so it can be called directly by other front-ends (a GUI/web
    app, a menu-driven script, a batch job, ...) without going through
    argparse. Returns a dict: {"rows", "lookup_ok", "lookup_unresolved",
    "pending", "dates_fixed", "sheet_load_errors"} -- everything the CLI
    used to just print, now handed back so a caller can render it however
    it wants (console text, a Streamlit success box, etc).
    """
    config = json.loads(read_text_robust(config_path))

    raw_wb = openpyxl.load_workbook(raw_path, data_only=True)
    raw_ws = raw_wb[raw_wb.sheetnames[0]] if not sheet else raw_wb[sheet]
    headers = [raw_ws.cell(row=1, column=c).value for c in range(1, raw_ws.max_column + 1)]

    rows = []
    for r in range(2, raw_ws.max_row + 1):
        row = {headers[c - 1]: raw_ws.cell(row=r, column=c).value
               for c in range(1, raw_ws.max_column + 1)}
        if any(v is not None for v in row.values()):
            rows.append(row)

    hedb_wb = openpyxl.load_workbook(hedb_path, data_only=True)
    lookup_maps, reverse_maps, sheet_load_errors = preload_lookup_maps(config, hedb_wb)

    # (sheet_load_errors, if any, is reported by the caller -- see cmd_build
    # below for the CLI's console-print version -- rather than printed here,
    # since this function is also called by non-console front-ends.)

    out = openpyxl.Workbook()
    sh = out.active
    sh.title = "Import_Ready"

    new_headers = [fdef["target_field"] for fdef in config["fields"]] + ["_NOTES"]
    sh.append(new_headers)
    for c in range(1, len(new_headers) + 1):
        cell = sh.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    stats = {"lookup_ok": 0, "lookup_unresolved": 0, "pending": 0, "dates_fixed": 0}
    date_cols = [i + 1 for i, fdef in enumerate(config["fields"]) if fdef.get("kind") == "date"]

    for row in rows:
        out_row = []
        notes = []
        for fdef in config["fields"]:
            raw_col = fdef["raw_column"]
            target = fdef["target_field"]
            kind = fdef.get("kind", "plain")
            v = row.get(raw_col)

            if raw_col not in row:
                notes.append(f"{target}: raw column '{raw_col}' not found in source sheet")

            # value_overrides: force specific raw values straight to a fixed
            # output, bypassing whatever the field's normal "kind" transform
            # would do -- e.g. a Lookup field where the HEDB sheet has no
            # entry for a special sentinel value like "INTL" (not a real
            # country), and the correct behavior is to leave it BLANK rather
            # than attempt (and fail) a lookup. Checked case-insensitively,
            # stripped of whitespace, before any other kind logic runs.
            overrides = fdef.get("value_overrides")
            override_matched = False
            if overrides and v not in (None, ""):
                vkey = str(v).strip().lower()
                for raw_k, out_v in overrides.items():
                    if str(raw_k).strip().lower() == vkey:
                        out_row.append(out_v)
                        notes.append(f"{target}: raw value '{v}' matched value_overrides -> forced to {out_v!r}")
                        override_matched = True
                        break
            if override_matched:
                continue

            # value_override_suffixes: same idea as value_overrides, but for
            # a whole family of sentinel values that share a trailing pattern
            # instead of one exact string -- e.g. a CIP code's generic
            # "Other/catch-all" sub-code within ANY family always ends in
            # ".9999" (99.9999 for "undeclared", but also 15.9999, 52.9999,
            # etc.), and none of those catch-all codes have their own row in
            # a reference list of specific named programs. Checked case-
            # insensitively, stripped of whitespace, before kind-specific logic.
            suffix_overrides = fdef.get("value_override_suffixes")
            suffix_matched = False
            if suffix_overrides and v not in (None, ""):
                vkey = str(v).strip().lower()
                for suffix, out_v in suffix_overrides.items():
                    if vkey.endswith(str(suffix).strip().lower()):
                        out_row.append(out_v)
                        notes.append(f"{target}: raw value '{v}' ends with '{suffix}' (value_override_suffixes) -> forced to {out_v!r}")
                        suffix_matched = True
                        break
            if suffix_matched:
                continue

            if kind == "lookup" and v not in (None, ""):
                sheet_name = fdef.get("hedb_sheet", "")
                key = _lookup_cache_key(fdef)
                direction = fdef.get("lookup_direction", "code_to_name")
                separator = fdef.get("combo_separator", " - ")
                combo_order = fdef.get("combo_order", "name_code")
                lut = lookup_maps.get(key, {})
                rev = reverse_maps.get(key, {})
                result = apply_lookup_transform(v, lut, rev, direction, separator, combo_order)
                if result is None:
                    reason = sheet_load_errors.get(key)
                    if reason:
                        notes.append(f"{target}: hedb_sheet not usable ({reason}) -- kept raw value as-is")
                    else:
                        notes.append(f"{target}: value '{v}' not found in '{sheet_name}' "
                                      f"(direction={direction}) -- kept as-is, verify manually")
                    out_row.append(v)
                    stats["lookup_unresolved"] += 1
                else:
                    out_row.append(result)
                    stats["lookup_ok"] += 1
            elif kind == "yesno":
                result, ok = apply_yesno_transform(
                    v, fdef.get("yes_label", "Yes"), fdef.get("no_label", "No")
                )
                out_row.append(result)
                if not ok:
                    notes.append(f"{target}: value '{v}' isn't a recognized Y/N variant -- kept as-is, verify manually")
                    stats["lookup_unresolved"] += 1
                else:
                    stats["lookup_ok"] += 1
            elif kind == "map":
                # For small, fixed dropdowns that AREN'T backed by any HEDB
                # sheet -- the choice list lives only inside the Zoho field
                # itself (e.g. a 2-3 option Dropdown like "FR-FRASCATI" /
                # "CR-CREATIVE ENDEAVORS"). "value_map" is written directly
                # in the config: {"raw_value": "exact Zoho choice text", ...}.
                # Matching is case-insensitive/stripped on the raw side.
                value_map = fdef.get("value_map", {})
                lookup_key = next(
                    (k for k in value_map if str(k).strip().lower() == str(v).strip().lower()),
                    None,
                )
                if lookup_key is not None:
                    out_row.append(value_map[lookup_key])
                    stats["lookup_ok"] += 1
                elif v in (None, ""):
                    out_row.append(v)
                else:
                    notes.append(f"{target}: value '{v}' not in value_map {list(value_map.keys())} -- kept as-is, verify manually")
                    out_row.append(v)
                    stats["lookup_unresolved"] += 1
            elif kind == "phone":
                out_row.append(apply_phone_transform(v))
            elif kind == "pending":
                out_row.append(v)
                note = fdef.get("note", "needs manual review before import")
                notes.append(f"{target}: {note}")
                stats["pending"] += 1
            elif kind == "date":
                # Zoho Date fields (as opposed to Date-Time fields) reject/mis-import
                # a value carrying a time component, e.g. "2024-11-01 00:00:00".
                # Strip the time and let the column-level number_format below
                # render it as a clean yyyy-mm-dd date.
                ok, parsed, _had_time = parse_date_value(v)
                if not ok:
                    notes.append(f"{target}: could not parse date value '{v}' -- check manually")
                    out_row.append(v)
                elif parsed is not None:
                    out_row.append(parsed)
                    stats["dates_fixed"] += 1
                else:
                    out_row.append(v)
            else:
                # plain field: tidy up the common "141.0 instead of 141" artifact
                # that comes from Excel storing whole numbers as floats.
                if isinstance(v, float) and v.is_integer():
                    v = int(v)
                out_row.append(v)

        out_row.append("; ".join(notes))
        sh.append(out_row)
        if date_cols:
            r_idx = sh.max_row
            for c_idx in date_cols:
                sh.cell(row=r_idx, column=c_idx).number_format = "yyyy-mm-dd"

    for c, h in enumerate(new_headers, start=1):
        sh.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max(16, min(45, len(h) + 4))
    sh.column_dimensions[openpyxl.utils.get_column_letter(len(new_headers))].width = 60
    sh.freeze_panes = "A2"

    out.save(out_path)
    stats["rows"] = len(rows)
    stats["sheet_load_errors"] = sheet_load_errors
    stats["out_path"] = out_path
    return stats


def cmd_build(args):
    stats = build_import_ready(args.raw, args.config, args.hedb, args.out, sheet=args.sheet)
    if stats["sheet_load_errors"]:
        print("WARNING -- these hedb_sheet values in your config are not usable yet "
              "(still a TODO placeholder, or the sheet name is wrong); affected "
              "columns will be left as raw codes with a note instead of stopping the run:")
        for lookup_key, err in stats["sheet_load_errors"].items():
            sheet_name = lookup_key[0] if isinstance(lookup_key, tuple) else lookup_key
            print(f"  - {sheet_name!r}: {err}")
    print(f"Wrote {stats['out_path']}  ({stats['rows']} rows)")
    print(f"Lookup fields resolved OK: {stats['lookup_ok']}")
    print(f"Lookup fields UNRESOLVED (check _NOTES): {stats['lookup_unresolved']}")
    print(f"Pending/custom-function fields (check _NOTES): {stats['pending']}")
    print(f"Date values cleaned (time stripped): {stats['dates_fixed']}")


# ---------------------------------------------------------------------------
# Step 3b: review -- annotate the CLIENT'S OWN raw sheet instead of building
# a converted one, so they can fix their own data before you (or they) run
# `build` for real. Same underlying lookup/map/yesno/phone/date logic as
# build_import_ready, but instead of writing the transformed value, it
# colors and comments the ORIGINAL cell to say what's wrong (or what will
# change) and leaves their file's layout untouched.
# ---------------------------------------------------------------------------

# Color legend -- kept as module-level constants so the CLI/Streamlit/README
# can all quote the same colors instead of three copies drifting apart.
REVIEW_COLORS = {
    "error":   {"fill": "FFC7CE", "font": "9C0006",
                "label": "ERROR -- must fix",
                "meaning": "Blocks a correct import: a value couldn't be understood at all "
                           "(e.g. an unparseable date), or the column needed for this field "
                           "is missing from the sheet entirely."},
    "warning": {"fill": "FFEB9C", "font": "9C6500",
                "label": "WARNING -- please check",
                "meaning": "The value doesn't match anything in the expected code/name list "
                           "or picklist -- it will be imported as-is, but it's very likely "
                           "wrong (typo, outdated code, wrong list) and needs a human look."},
    "info":    {"fill": "BDD7EE", "font": "1F4E78",
                "label": "INFO -- will be auto-converted",
                "meaning": "Nothing to fix -- just a heads-up that this value will look "
                           "different after the real import-ready sheet is built (e.g. a "
                           "code becomes its descriptive name, a date's time portion is "
                           "dropped, a phone number's punctuation is stripped)."},
}


def classify_cell(fdef, v, lookup_maps, reverse_maps, sheet_load_errors):
    """
    Look at one raw cell's value against its field definition and decide:
    does this need to change before import? Returns (severity, message)
    where severity is one of "error"/"warning"/"info"/None (None = no issue).
    This mirrors build_import_ready's per-kind logic exactly, but classifies
    instead of transforming -- the two should always be updated together.
    """
    if v in (None, ""):
        return None, None

    kind = fdef.get("kind", "plain")
    overrides = fdef.get("value_overrides")
    if overrides:
        vkey = str(v).strip().lower()
        for raw_k, out_v in overrides.items():
            if str(raw_k).strip().lower() == vkey:
                shown = "(blank)" if out_v == "" else f"'{out_v}'"
                return "info", f"Recognized special value -- will be set to {shown} instead of looked up."

    suffix_overrides = fdef.get("value_override_suffixes")
    if suffix_overrides:
        vkey = str(v).strip().lower()
        for suffix, out_v in suffix_overrides.items():
            if vkey.endswith(str(suffix).strip().lower()):
                shown = "(blank)" if out_v == "" else f"'{out_v}'"
                return "info", f"Ends with '{suffix}' -- recognized as a catch-all/special value, will be set to {shown} instead of looked up."

    if kind == "lookup":
        sheet_name = fdef.get("hedb_sheet", "")
        key = _lookup_cache_key(fdef)
        if key in sheet_load_errors:
            return "error", (f"Can't verify this value yet -- the reference list "
                              f"'{sheet_name}' isn't set up ({sheet_load_errors[key]}).")
        direction = fdef.get("lookup_direction", "code_to_name")
        separator = fdef.get("combo_separator", " - ")
        combo_order = fdef.get("combo_order", "name_code")
        lut = lookup_maps.get(key, {})
        rev = reverse_maps.get(key, {})
        result = apply_lookup_transform(v, lut, rev, direction, separator, combo_order)
        if result is None:
            return "warning", (f"'{v}' was not found in the '{sheet_name}' reference list -- "
                                f"check this code/name is correct.")
        if str(result).strip() == str(v).strip():
            return None, None
        return "info", f"Will be converted to '{result}'."

    if kind == "map":
        value_map = fdef.get("value_map", {})
        key = next((k for k in value_map if str(k).strip().lower() == str(v).strip().lower()), None)
        if key is None:
            return "warning", (f"'{v}' doesn't match any of this field's expected values -- "
                                f"check it against the accepted list.")
        result = value_map[key]
        if str(result).strip() == str(v).strip():
            return None, None
        return "info", f"Will be converted to '{result}'."

    if kind == "yesno":
        key = str(v).strip().lower()
        if key in YES_VALUES or key in NO_VALUES:
            label = fdef.get("yes_label", "Yes") if key in YES_VALUES else fdef.get("no_label", "No")
            if str(v).strip() == label:
                return None, None
            return "info", f"Will be converted to '{label}'."
        return "warning", f"'{v}' isn't a recognized Yes/No value (expected Y/N/Yes/No/1/0/True/False)."

    if kind == "phone":
        cleaned = apply_phone_transform(v)
        if cleaned in (None, "", "+"):
            return "warning", f"'{v}' doesn't contain any digits -- doesn't look like a valid phone number."
        if str(cleaned) == str(v).strip():
            return None, None
        return "info", f"Will be reformatted to '{cleaned}' (digits + leading + only)."

    if kind == "date":
        ok, parsed, had_time = parse_date_value(v)
        if not ok:
            return "error", f"'{v}' doesn't look like a valid date -- please use YYYY-MM-DD."
        if had_time:
            return "info", "Has a time portion that will be stripped -- only the date part is kept."
        return None, None

    if kind == "pending":
        note = fdef.get("note", "needs manual review before import")
        return "warning", f"Needs manual review: {note}"

    return None, None


def describe_field_transform(fdef):
    """
    One-line, plain-language description of what happens to a field's value
    on the way into the real import-ready sheet -- reused for the mapping
    config's own "_note", and for the review report's header comments/
    Column Mapping sheet, so the wording never drifts between the two.
    Falls back to this generated text only if the config doesn't already
    carry a "_note" (existing configs already have one from an earlier pass).
    """
    if fdef.get("_note"):
        return fdef["_note"]

    kind = fdef.get("kind", "plain")

    if kind == "plain":
        return "No conversion -- value stays exactly the same."

    if kind == "date":
        return ("Date-time values are trimmed to a plain date (time portion removed) and "
                "written as yyyy-mm-dd.")

    if kind == "phone":
        return ("Phone number is cleaned to digits only (plus a leading '+' if present) -- "
                "dashes/spaces/punctuation are stripped.")

    if kind == "yesno":
        yes_label = fdef.get("yes_label", "Yes")
        no_label = fdef.get("no_label", "No")
        return f"Raw Y/N (or true/false, 1/0) is converted to '{yes_label}'/'{no_label}'."

    if kind == "map":
        vm = fdef.get("value_map", {})
        items = list(vm.items())
        if len(items) <= 4:
            pairs = "; ".join(f"'{k}'->'{v}'" for k, v in items)
            return f"Converted to the exact Zoho picklist text: {pairs}."
        return f"Converted to the exact Zoho picklist text using a {len(items)}-entry table."

    if kind == "lookup":
        sheet_name = fdef.get("hedb_sheet", "?")
        direction = fdef.get("lookup_direction", "code_to_name")
        if direction == "code_to_name":
            return f"Code converted to its descriptive Name using the '{sheet_name}' list."
        if direction == "name_to_code":
            return f"Descriptive Name converted to its code using the '{sheet_name}' list."
        return f"Converted to a combined Name/Code display value using the '{sheet_name}' list."

    if kind == "pending":
        return "Needs manual review -- " + fdef.get("note", "custom logic, unverified.")

    return "No conversion -- value stays exactly the same."


SEVERITY_LABELS = {
    "error":   "MUST FIX",
    "warning": "PLEASE CHECK",
    "info":    "INFO (no action needed)",
}


def build_issue_report(raw_path, config_path, hedb_path, out_path, sheet=None):
    """
    Build a pure REPORT about the client's raw sheet -- it never touches,
    copies, or reproduces their actual data. Two tables only:
      1. "Column Changes"  -- one row per mapped field: does this raw column
         get renamed, is it missing entirely, and what happens to its value
         (including "stays exactly the same" for plain fields).
      2. "Row Issues"       -- one row per flagged CELL, citing the Excel row
         number, the raw column name, the current value, and what to do --
         so the client can jump straight to that row in their own file.
    Nothing here is colored/annotated on a copy of their sheet; this is
    strictly a findings report to hand them so THEY make the edits.
    Returns a dict: {"rows", "errors", "warnings", "infos", "missing_columns"}.
    """
    config = json.loads(read_text_robust(config_path))

    raw_wb = openpyxl.load_workbook(raw_path, data_only=True)
    raw_ws = raw_wb[raw_wb.sheetnames[0]] if not sheet else raw_wb[sheet]
    headers = [raw_ws.cell(row=1, column=c).value for c in range(1, raw_ws.max_column + 1)]
    header_to_col = {h: c + 1 for c, h in enumerate(headers) if h}

    hedb_wb = openpyxl.load_workbook(hedb_path, data_only=True)
    lookup_maps, reverse_maps, sheet_load_errors = preload_lookup_maps(config, hedb_wb)

    field_to_col = {}
    missing_columns = []
    for fdef in config["fields"]:
        rc = fdef["raw_column"]
        if rc in header_to_col:
            field_to_col[fdef["target_field"]] = header_to_col[rc]
        else:
            missing_columns.append({"target_field": fdef["target_field"], "raw_column": rc})

    out = openpyxl.Workbook()

    # --- Sheet 1: Column Changes ---------------------------------------
    colmap = out.active
    colmap.title = "Column Changes"
    colmap_headers = ["#", "Raw column (in your file)", "Becomes Zoho field", "Status", "What happens to the value"]
    for c, h in enumerate(colmap_headers, start=1):
        cell = colmap.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    r = 2
    for i, fdef in enumerate(config["fields"], start=1):
        raw_col = fdef["raw_column"]
        target = fdef["target_field"]
        found = raw_col in header_to_col
        colmap.cell(row=r, column=1, value=i)
        colmap.cell(row=r, column=2, value=raw_col)
        colmap.cell(row=r, column=3, value=target)
        if not found:
            status, detail = "MISSING", "Column not found in your file -- can't be checked or imported until it exists."
        elif raw_col == target:
            status, detail = "No rename needed", describe_field_transform(fdef)
        else:
            status, detail = "Rename required", describe_field_transform(fdef)
        status_cell = colmap.cell(row=r, column=4, value=status)
        colmap.cell(row=r, column=5, value=detail).alignment = Alignment(wrap_text=True)
        if status == "MISSING":
            status_cell.font = Font(color=REVIEW_COLORS["error"]["font"], bold=True)
            status_cell.fill = PatternFill("solid", fgColor=REVIEW_COLORS["error"]["fill"])
        elif status == "Rename required":
            status_cell.font = Font(color=REVIEW_COLORS["warning"]["font"], bold=True)
            status_cell.fill = PatternFill("solid", fgColor=REVIEW_COLORS["warning"]["fill"])
        r += 1
    for c, w in enumerate((5, 30, 30, 18, 62), start=1):
        colmap.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    colmap.freeze_panes = "A2"
    colmap.auto_filter.ref = f"A1:E{r - 1}"

    # --- Sheet 2: Row Issues --------------------------------------------
    issues = out.create_sheet("Row Issues")
    issue_headers = ["Row # (in your file)", "Raw column", "Becomes Zoho field", "Current value", "Type", "What to check / do"]
    for c, h in enumerate(issue_headers, start=1):
        cell = issues.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    counts = {"error": 0, "warning": 0, "info": 0}
    n_rows = 0
    max_row = raw_ws.max_row
    max_col = max(len(headers), 1)
    r = 2

    for row_num in range(2, max_row + 1):
        row_values = [raw_ws.cell(row=row_num, column=c).value for c in range(1, max_col + 1)]
        if all(v is None for v in row_values):
            continue
        n_rows += 1

        for fdef in config["fields"]:
            col = field_to_col.get(fdef["target_field"])
            if col is None:
                continue
            v = raw_ws.cell(row=row_num, column=col).value
            sev, msg = classify_cell(fdef, v, lookup_maps, reverse_maps, sheet_load_errors)
            if sev is None:
                continue
            counts[sev] += 1
            issues.cell(row=r, column=1, value=row_num)
            issues.cell(row=r, column=2, value=fdef["raw_column"])
            issues.cell(row=r, column=3, value=fdef["target_field"])
            issues.cell(row=r, column=4, value=v)
            type_cell = issues.cell(row=r, column=5, value=SEVERITY_LABELS[sev])
            type_cell.font = Font(color=REVIEW_COLORS[sev]["font"], bold=True)
            type_cell.fill = PatternFill("solid", fgColor=REVIEW_COLORS[sev]["fill"])
            issues.cell(row=r, column=6, value=msg or "").alignment = Alignment(wrap_text=True)
            r += 1

    if r == 2:
        issues.cell(row=2, column=1, value="No issues found -- every checkable value matched its expected list/format.")

    for c, w in enumerate((14, 26, 26, 20, 22, 55), start=1):
        issues.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    issues.freeze_panes = "A2"
    if r > 2:
        issues.auto_filter.ref = f"A1:F{r - 1}"

    out.save(out_path)
    return {
        "rows": n_rows,
        "errors": counts["error"],
        "warnings": counts["warning"],
        "infos": counts["info"],
        "missing_columns": missing_columns,
        "out_path": out_path,
    }


def cmd_report(args):
    stats = build_issue_report(args.raw, args.config, args.hedb, args.out, sheet=args.sheet)
    print(f"Wrote {stats['out_path']}  ({stats['rows']} rows checked)")
    print(f"Must fix:      {stats['errors']}")
    print(f"Please check:  {stats['warnings']}")
    print(f"Info only:     {stats['infos']}")
    if stats["missing_columns"]:
        print("\nColumns expected but not found in this file at all:")
        for mc in stats["missing_columns"]:
            print(f"  - expected header {mc['raw_column']!r} (for field {mc['target_field']!r})")
    print("\nOpen the report: 'Column Changes' sheet lists every column and whether it's "
          "renamed/missing; 'Row Issues' sheet lists the specific rows/cells to fix.")


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="command", required=True)

    p1 = sub.add_parser("analyze-script", help="Extract field usage/classification from a Deluge workflow script")
    p1.add_argument("script", help="Path to a .txt file containing the form's Deluge add/update script")
    p1.add_argument("--out", default=None,
                    help="Write the starter config directly to this .json path (UTF-8). "
                         "Recommended on Windows/PowerShell instead of `> file.json`, "
                         "which defaults to UTF-16 and breaks the JSON parser later.")
    p1.set_defaults(func=cmd_analyze_script)

    pA = sub.add_parser("auto-map", help="Auto-generate a mapping config from a script + your raw data + HEDB_.xlsx")
    pA.add_argument("script", help="Path to a .txt file containing the form's Deluge add/update script")
    pA.add_argument("--raw", required=True, help="Path to the raw data .xlsx (used to fuzzy-match column headers and sample values)")
    pA.add_argument("--sheet", default=None, help="Sheet name in --raw (default: first sheet)")
    pA.add_argument("--hedb", required=True, help="Path to HEDB_.xlsx (used to auto-detect lookup sheets)")
    pA.add_argument("--out", default=None, help="Write the config directly to this .json path (UTF-8)")
    pA.set_defaults(func=cmd_auto_map)

    pD = sub.add_parser("auto-map-ds", help="Most accurate auto-generate: build a config straight from a .ds export's field definitions + push workflow + your raw data")
    pD.add_argument("ds", help="Path to the whole-application .ds export file")
    pD.add_argument("form", help="Form name exactly as it appears in `list-forms` (e.g. Students_SOD_Applicants)")
    pD.add_argument("--raw", required=True, help="Path to the raw data .xlsx (used to match column headers)")
    pD.add_argument("--sheet", default=None, help="Sheet name in --raw (default: first sheet)")
    pD.add_argument("--hedb", default=None, help="Path to HEDB_.xlsx (optional -- used to sanity-check that guessed hedb_sheet names actually exist)")
    pD.add_argument("--workflow-var", default=None, help="Which workflow block to treat as authoritative when a form has more than one (e.g. staging vs production) -- pass the variable name shown in the command's own output. Default: prefers one with 'production' in its name/label, else the last one found.")
    pD.add_argument("--out", default=None, help="Write the config directly to this .json path (UTF-8)")
    pD.set_defaults(func=cmd_auto_map_ds)

    pL = sub.add_parser("list-forms", help="List every form name in a whole-application .ds export")
    pL.add_argument("ds", help="Path to the .ds application export file")
    pL.set_defaults(func=cmd_list_forms)

    pE = sub.add_parser("extract-form", help="Pull one form's field definitions + push workflow(s) out of a .ds export")
    pE.add_argument("ds", help="Path to the .ds application export file")
    pE.add_argument("form", help="Form name exactly as it appears in `list-forms` (e.g. Student_Research)")
    pE.add_argument("--out-prefix", default=None, help="Filename prefix for the extracted .form/.txt files (default: the form name)")
    pE.set_defaults(func=cmd_extract_form)

    p2 = sub.add_parser("describe-sheet", help="Dump a HEDB_.xlsx sheet's rows to find code/name columns")
    p2.add_argument("hedb", help="Path to HEDB_.xlsx")
    p2.add_argument("sheet", help="Exact sheet name (or a fragment to get suggestions)")
    p2.add_argument("--rows", type=int, default=10)
    p2.set_defaults(func=cmd_describe_sheet)

    p3 = sub.add_parser("build", help="Transform a raw data sheet into an import-ready sheet")
    p3.add_argument("--raw", required=True, help="Path to the raw data .xlsx to correct")
    p3.add_argument("--sheet", default=None, help="Sheet name in --raw (default: first sheet)")
    p3.add_argument("--config", required=True, help="Path to the mapping-config .json (see analyze-script)")
    p3.add_argument("--hedb", required=True, help="Path to HEDB_.xlsx")
    p3.add_argument("--out", required=True, help="Output .xlsx path")
    p3.set_defaults(func=cmd_build)

    p4 = sub.add_parser("report", help="Produce a pure FINDINGS REPORT about the client's raw sheet (never touches/copies their data) -- for when you want them to fix their own file")
    p4.add_argument("--raw", required=True, help="Path to the raw data .xlsx to check")
    p4.add_argument("--sheet", default=None, help="Sheet name in --raw (default: first sheet)")
    p4.add_argument("--config", required=True, help="Path to the mapping-config .json")
    p4.add_argument("--hedb", required=True, help="Path to HEDB_.xlsx")
    p4.add_argument("--out", required=True, help="Output .xlsx path (a 'Column Changes' sheet plus a 'Row Issues' sheet)")
    p4.set_defaults(func=cmd_report)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
